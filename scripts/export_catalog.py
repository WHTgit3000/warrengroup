import json
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\damia\Downloads\Warren Group Final Working Sheet - 4.15.26 (3).xlsx")
OUTPUT_PATH = Path(__file__).resolve().parents[1] / "data" / "catalog.json"


def clean_price(value):
    if value in (None, "", "$ -"):
        return None
    if isinstance(value, str):
        value = value.replace("$", "").strip()
        if value == "-":
            return None
    return float(value)


def title_case_county(value):
    return str(value or "").strip().title()


def build_item(state, county, category, price, link):
    county_name = title_case_county(county)
    product_name = f"{category} - {county_name} County, {state}"
    slug = f"{category.lower()}-{state.lower()}-{county_name.lower().replace(' ', '-')}"
    return {
        "id": slug,
        "productName": product_name,
        "category": category,
        "state": state,
        "county": county_name,
        "basePrice": price,
        "link": link,
    }


def build_request_option(state, county, category):
    county_name = title_case_county(county)
    slug = f"{category.lower()}-{state.lower()}-{county_name.lower().replace(' ', '-')}"
    return {
        "id": slug,
        "state": state,
        "county": county_name,
        "category": category,
        "label": f"{county_name} County, {state}",
    }


def main():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["RAW_Data"]
    items = []
    request_options = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
      state, _, county, _, divorce_available, divorce_price, divorce_link, probate_available, probate_price, probate_link = row

      if county and divorce_available == "Y":
          price = clean_price(divorce_price)
          if price and divorce_link:
              items.append(build_item(state, county, "Divorce", price, divorce_link))
      elif county:
          request_options.append(build_request_option(state, county, "Divorce"))

      if county and probate_available == "Y":
          price = clean_price(probate_price)
          if price and probate_link:
              items.append(build_item(state, county, "Probate", price, probate_link))
      elif county:
          request_options.append(build_request_option(state, county, "Probate"))

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(
            {
                "generatedAt": datetime.now(UTC).isoformat(),
                "sourceWorkbook": str(WORKBOOK_PATH),
                "items": items,
                "requestOptions": request_options,
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"Wrote {len(items)} items and {len(request_options)} request options to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
